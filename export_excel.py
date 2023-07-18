import json
import openpyxl
from urllib.parse import urlparse, parse_qs, unquote
from sqlalchemy import create_engine, text

engine = create_engine("sqlite:///mydb.sql", echo=True)


def extract_q_lang_from_url(url):
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)

    split_params = str(query_params.get('q')[0]).split(sep=' ')

    keyword_value = split_params[0]
    language_value = split_params[1].split(sep=':')[-1]

    return keyword_value, language_value


with engine.connect() as connection:
    tweet_query = connection.execute(text('''
        SELECT
            Query_URL,
            Profile_URL,
            Tweet_URL,
            strftime('%Y-%m-%d %H:%M:%S', Date_Time) AS Tweet_Datetime,
            Likes_Count,
            Reply_Count,
            Retweet_Count,
            Tweet_Text
        FROM tweet_info 
        WHERE Tweet_Text IS NOT NULL AND Tweet_Datetime IS NOT NULL
        ORDER BY Tweet_Datetime DESC;

    '''))

    # Fetch all the data from the CursorResult object
    data = tweet_query.fetchall()

    # Create an Excel workbook and a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    data_sets = list()

    column_names = ['Keyword', 'Language', 'Profile URL', 'Tweet URL', 'Datetime', 'Likes', 'Reply', 'Retweet', 'Text']

    # Write the column names to the first row of the Excel worksheet
    for column_index, column_name in enumerate(column_names):
        worksheet.cell(row=1, column=column_index + 1, value=column_name)

    # Write the data to the Excel worksheet
    for row_index, row_data in enumerate(data):
        keyword, language = extract_q_lang_from_url(url=row_data[0])

        removed_first_element = row_data[1:]
        _row_data = (keyword, language) + removed_first_element

        single_element = {
            "keyword": _row_data[0],
            "language": _row_data[1],
            "profile_url": _row_data[2],
            "tweet_url": _row_data[3],
            "datetime": _row_data[4],
            "likes": _row_data[5],
            "reply": _row_data[6],
            "retweet": _row_data[7],
            "text": _row_data[8]
        }

        data_sets.append(single_element)

        for column_index, cell_value in enumerate(_row_data):
            worksheet.cell(row=row_index + 2, column=column_index + 1, value=cell_value)

    # Save the Excel workbook to a file
    workbook.save("output.xlsx")

all_the_data = {"data": data_sets}

json_data = json.dumps(all_the_data, indent=4)

with open("output.json", "w") as write_json:
    write_json.write(json_data)
    write_json.close()
